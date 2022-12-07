import openpyxl
from pathlib import Path

root_dir = Path(__file__).parent


def get_data(sheet_name, test_scenario):
    workbook = openpyxl.load_workbook(root_dir/"TestData/td_meralco_online.xlsx")
    sheet = workbook[f"{sheet_name}"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):
        if sheet.cell(row=i, column=1).value == test_scenario:
            dataList = []
            for j in range(1, totalcols + 1):
                data = sheet.cell(row=i, column=j).value
                dataList.insert(j, data)
            mainList.insert(i, dataList)
            print(mainList)
    return mainList
